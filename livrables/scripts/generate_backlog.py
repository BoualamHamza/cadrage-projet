"""Génère Backlog_Fashion_Insta.xlsx à partir du template fourni.

Contient :
- Onglet 'Backlog' : user stories IA priorisées MoSCoW + chiffrage en jours par profil
- Onglet 'Synthèse coûts' : total jours et coûts par profil
- Onglet 'Planning sprints' : répartition des US par sprint
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
TEMPLATE = ROOT / "pieces_jointes" / "Template_Backlog.xlsx"
OUT = ROOT / "livrables" / "Backlog_Fashion_Insta.xlsx"

# ---------- Profils & TJM ----------
PROFILS = [
    ("PO / Product Manager", 600),
    ("Tech Lead", 800),
    ("Data Scientist junior", 500),
    ("Data Scientist senior (sous-traitant)", 1000),
    ("ML / Computer Vision Engineer", 700),
    ("Développeur mobile", 600),
    ("DevOps / Cloud Azure", 700),
    ("UX / UI Designer", 550),
    ("QA / Testeur", 500),
]

# ---------- Backlog : 15 User Stories ----------
# Colonnes chiffrage : pour chaque US, charge totale (j) + % par profil
# (PM, TL, DSJ, DSS, MLE, MOB, OPS, UX, QA)
US = [
    # id, titre, en_tant_que, je_veux, afin_de, points, donnees, moscow, sprint, mvp, charge_j, repartition_%
    {
        "id": "US01",
        "titre": "Création de compte / connexion sécurisée",
        "tant_que": "Utilisateur de l'application",
        "veut": "créer un compte et me connecter avec adresse mail + mot de passe",
        "afin": "accéder à mon espace personnel et mes recommandations",
        "points": 3,
        "donnees": "- adresse mail\n- mot de passe (haché)\n- token de session",
        "moscow": "MUST",
        "sprint": 1,
        "mvp": "Oui",
        "charge": 8,
        "rep": {"PO / Product Manager": 10, "Tech Lead": 10, "Développeur mobile": 50, "DevOps / Cloud Azure": 10, "UX / UI Designer": 10, "QA / Testeur": 10},
    },
    {
        "id": "US02",
        "titre": "Prise de photo et alimentation de la garde-robe",
        "tant_que": "Utilisateur",
        "veut": "prendre une photo de moi avec un vêtement et l'ajouter à ma garde-robe",
        "afin": "constituer une collection de photos analysables par l'IA",
        "points": 8,
        "donnees": "- photo utilisateur (image)\n- métadonnées (date, type de vêtement)\n- identifiant utilisateur",
        "moscow": "MUST",
        "sprint": 1,
        "mvp": "Oui",
        "charge": 12,
        "rep": {"PO / Product Manager": 5, "Tech Lead": 5, "Développeur mobile": 55, "DevOps / Cloud Azure": 10, "UX / UI Designer": 15, "QA / Testeur": 10},
    },
    {
        "id": "US03",
        "titre": "Gestion de la collection de photos (visualisation, suppression)",
        "tant_que": "Utilisateur",
        "veut": "visualiser, organiser et supprimer les photos de ma garde-robe",
        "afin": "garder le contrôle sur les données utilisées par l'algorithme",
        "points": 5,
        "donnees": "- liste de photos\n- métadonnées",
        "moscow": "MUST",
        "sprint": 2,
        "mvp": "Oui",
        "charge": 8,
        "rep": {"PO / Product Manager": 5, "Développeur mobile": 65, "UX / UI Designer": 15, "QA / Testeur": 15},
    },
    {
        "id": "US04",
        "titre": "[IA] Détection / segmentation du vêtement sur la photo",
        "tant_que": "Système IA",
        "veut": "détecter et segmenter automatiquement le vêtement principal sur les photos utilisateur",
        "afin": "construire un vecteur de représentation du style pour la recommandation",
        "points": 20,
        "donnees": "- photos utilisateur\n- jeu d'entraînement annoté (DeepFashion ou équivalent)\n- masques de segmentation",
        "moscow": "MUST",
        "sprint": 2,
        "mvp": "Oui",
        "charge": 30,
        "rep": {"PO / Product Manager": 3, "Tech Lead": 7, "Data Scientist junior": 25, "Data Scientist senior (sous-traitant)": 30, "ML / Computer Vision Engineer": 25, "DevOps / Cloud Azure": 5, "QA / Testeur": 5},
    },
    {
        "id": "US05",
        "titre": "[IA] Recommandation d'articles à partir de la garde-robe",
        "tant_que": "Utilisateur",
        "veut": "recevoir des suggestions d'articles du catalogue alignées avec le style de mes photos",
        "afin": "découvrir des vêtements correspondant à mes goûts",
        "points": 20,
        "donnees": "- embeddings garde-robe\n- catalogue produits + images\n- historique de feedback",
        "moscow": "MUST",
        "sprint": 3,
        "mvp": "Oui",
        "charge": 25,
        "rep": {"PO / Product Manager": 5, "Tech Lead": 10, "Data Scientist junior": 25, "Data Scientist senior (sous-traitant)": 25, "ML / Computer Vision Engineer": 20, "DevOps / Cloud Azure": 10, "QA / Testeur": 5},
    },
    {
        "id": "US06",
        "titre": "Affichage du vêtement recommandé sur la photo utilisateur (try-on)",
        "tant_que": "Utilisateur",
        "veut": "voir le vêtement recommandé incrusté sur ma photo",
        "afin": "juger visuellement avant d'acheter",
        "points": 13,
        "donnees": "- photo utilisateur\n- image article (texture, masque)\n- résultat composé",
        "moscow": "SHOULD",
        "sprint": 4,
        "mvp": "Non",
        "charge": 18,
        "rep": {"PO / Product Manager": 5, "Tech Lead": 5, "Data Scientist senior (sous-traitant)": 25, "ML / Computer Vision Engineer": 40, "Développeur mobile": 15, "UX / UI Designer": 5, "QA / Testeur": 5},
    },
    {
        "id": "US07",
        "titre": "Personnalisation : couleur / style du vêtement recommandé",
        "tant_que": "Utilisateur",
        "veut": "modifier la couleur (et éventuellement le style — manches courtes/longues) du vêtement proposé",
        "afin": "affiner la recommandation à mes envies du moment",
        "points": 8,
        "donnees": "- attributs produit (couleur, style)\n- variantes catalogue",
        "moscow": "COULD",
        "sprint": 5,
        "mvp": "Non",
        "charge": 10,
        "rep": {"PO / Product Manager": 10, "ML / Computer Vision Engineer": 25, "Développeur mobile": 45, "UX / UI Designer": 10, "QA / Testeur": 10},
    },
    {
        "id": "US08",
        "titre": "Définition des styles préférés",
        "tant_que": "Utilisateur",
        "veut": "sélectionner mes styles vestimentaires préférés parmi une liste",
        "afin": "orienter les recommandations indépendamment de mes photos",
        "points": 3,
        "donnees": "- catalogue de styles\n- préférences utilisateur",
        "moscow": "MUST",
        "sprint": 2,
        "mvp": "Oui",
        "charge": 5,
        "rep": {"PO / Product Manager": 10, "Développeur mobile": 50, "UX / UI Designer": 20, "QA / Testeur": 20},
    },
    {
        "id": "US09",
        "titre": "Référencement marques préférées",
        "tant_que": "Utilisateur",
        "veut": "indiquer mes marques préférées du groupe Fashion-Insta",
        "afin": "filtrer les recommandations selon mes marques favorites",
        "points": 3,
        "donnees": "- liste des marques\n- préférences utilisateur",
        "moscow": "SHOULD",
        "sprint": 3,
        "mvp": "Non",
        "charge": 4,
        "rep": {"PO / Product Manager": 10, "Développeur mobile": 60, "UX / UI Designer": 15, "QA / Testeur": 15},
    },
    {
        "id": "US10",
        "titre": "Référencement sources d'inspiration (blogs / influenceurs)",
        "tant_que": "Utilisateur",
        "veut": "ajouter des blogs et influenceurs comme sources d'inspiration",
        "afin": "enrichir l'IA avec mes références personnelles",
        "points": 8,
        "donnees": "- URL/handles sources\n- crawl ou API tierce (avec consentement)",
        "moscow": "COULD",
        "sprint": 6,
        "mvp": "Non",
        "charge": 10,
        "rep": {"PO / Product Manager": 10, "Data Scientist junior": 30, "ML / Computer Vision Engineer": 20, "Développeur mobile": 25, "DevOps / Cloud Azure": 5, "QA / Testeur": 10},
    },
    {
        "id": "US11",
        "titre": "[IA] Recommandation à partir des préférences et tendances",
        "tant_que": "Utilisateur",
        "veut": "recevoir des recommandations basées sur mes préférences (styles, marques) et les tendances",
        "afin": "découvrir des articles alignés avec mes goûts déclarés",
        "points": 13,
        "donnees": "- préférences utilisateur\n- tendances catalogue\n- historique d'achats agrégé (anonymisé)",
        "moscow": "MUST",
        "sprint": 3,
        "mvp": "Oui",
        "charge": 18,
        "rep": {"PO / Product Manager": 5, "Tech Lead": 5, "Data Scientist junior": 30, "Data Scientist senior (sous-traitant)": 25, "ML / Computer Vision Engineer": 20, "DevOps / Cloud Azure": 10, "QA / Testeur": 5},
    },
    {
        "id": "US12",
        "titre": "Avis / feedback sur les recommandations",
        "tant_que": "Utilisateur",
        "veut": "noter (j'aime / je n'aime pas) les recommandations qui me sont faites",
        "afin": "affiner le modèle au fil du temps",
        "points": 5,
        "donnees": "- identifiant utilisateur\n- identifiant article\n- score / label",
        "moscow": "SHOULD",
        "sprint": 4,
        "mvp": "Non",
        "charge": 7,
        "rep": {"PO / Product Manager": 10, "Data Scientist junior": 25, "Développeur mobile": 45, "UX / UI Designer": 10, "QA / Testeur": 10},
    },
    {
        "id": "US13",
        "titre": "[RGPD] Consultation / modification / suppression des données personnelles",
        "tant_que": "Utilisateur",
        "veut": "accéder à mes données, les modifier et les supprimer",
        "afin": "exercer mes droits RGPD",
        "points": 8,
        "donnees": "- toutes les données personnelles stockées\n- logs d'accès",
        "moscow": "MUST",
        "sprint": 4,
        "mvp": "Oui",
        "charge": 10,
        "rep": {"PO / Product Manager": 10, "Tech Lead": 10, "Développeur mobile": 50, "DevOps / Cloud Azure": 15, "QA / Testeur": 15},
    },
    {
        "id": "US14",
        "titre": "[RGPD] Désinscription service recommandation + purge automatique",
        "tant_que": "Utilisateur",
        "veut": "me désinscrire du service de recommandation et déclencher la suppression de mes photos",
        "afin": "ne plus être profilé et faire effacer mes données sensibles",
        "points": 5,
        "donnees": "- consentement utilisateur\n- politique de rétention",
        "moscow": "MUST",
        "sprint": 4,
        "mvp": "Oui",
        "charge": 8,
        "rep": {"PO / Product Manager": 10, "Tech Lead": 10, "Développeur mobile": 40, "DevOps / Cloud Azure": 25, "QA / Testeur": 15},
    },
    {
        "id": "US15",
        "titre": "[RGPD] Purge automatique après inactivité prolongée",
        "tant_que": "Système",
        "veut": "détecter l'inactivité et purger automatiquement les données sensibles après le délai défini",
        "afin": "respecter la limitation de conservation prévue par le RGPD",
        "points": 5,
        "donnees": "- dernière activité utilisateur\n- politique de rétention",
        "moscow": "MUST",
        "sprint": 5,
        "mvp": "Oui",
        "charge": 6,
        "rep": {"Tech Lead": 15, "DevOps / Cloud Azure": 55, "Développeur mobile": 15, "QA / Testeur": 15},
    },
    {
        "id": "US16",
        "titre": "[Commande] Ajout au panier et validation de la commande",
        "tant_que": "Utilisateur",
        "veut": "ajouter un article recommandé à mon panier et valider ma commande sans quitter l'application",
        "afin": "transformer une recommandation en achat effectif, sans rupture de parcours vers le site web",
        "points": 8,
        "donnees": "- identifiant utilisateur\n- panier (articles, quantités)\n- API panier/commande du site e-commerce existant",
        "moscow": "MUST",
        "sprint": 3,
        "mvp": "Oui",
        "charge": 7,
        "rep": {"PO / Product Manager": 10, "Tech Lead": 10, "Développeur mobile": 55, "DevOps / Cloud Azure": 10, "QA / Testeur": 15},
    },
    {
        "id": "US17",
        "titre": "[Commande] Choix du mode de livraison et paiement",
        "tant_que": "Utilisateur",
        "veut": "choisir un mode de livraison et régler ma commande par carte bancaire",
        "afin": "finaliser mon achat de bout en bout depuis l'application mobile",
        "points": 8,
        "donnees": "- adresse de livraison\n- mode de livraison\n- moyen de paiement (tokenisé, via le prestataire de paiement existant)",
        "moscow": "MUST",
        "sprint": 3,
        "mvp": "Oui",
        "charge": 6,
        "rep": {"PO / Product Manager": 10, "Tech Lead": 15, "Développeur mobile": 45, "DevOps / Cloud Azure": 20, "QA / Testeur": 10},
    },
]

# ---------- Styles ----------
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
MOSCOW_COLORS = {"MUST": "C00000", "SHOULD": "ED7D31", "COULD": "FFC000", "WONT": "808080"}
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def build():
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb["Backlog"]

    # Réécriture du header (le template a 6 colonnes ; on en met davantage)
    headers = [
        "ID", "Titre", "En tant que", "Je veux...", "...afin de...",
        "Story Points", "MoSCoW", "Sprint", "MVP",
        "Données nécessaires",
        "Charge totale (j)",
    ] + [p[0] for p in PROFILS] + [
        "Coût US (€)",
    ]
    # purge existing rows
    ws.delete_rows(1, ws.max_row)

    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 40

    # Données
    for r, us in enumerate(US, start=2):
        charge = us["charge"]
        rep = us["rep"]
        # jours par profil
        jours_par_profil = []
        cost_total = 0.0
        for prof, tjm in PROFILS:
            pct = rep.get(prof, 0)
            j = round(charge * pct / 100, 2)
            jours_par_profil.append(j)
            cost_total += j * tjm

        row_vals = [
            us["id"], us["titre"], us["tant_que"], us["veut"], us["afin"],
            us["points"], us["moscow"], us["sprint"], us["mvp"],
            us["donnees"], charge,
        ] + jours_par_profil + [round(cost_total, 0)]

        for c, v in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDER
            cell.alignment = LEFT_WRAP if c in (2, 3, 4, 5, 10) else CENTER

        # couleur MoSCoW
        mc = ws.cell(row=r, column=7)
        mc.fill = PatternFill("solid", fgColor=MOSCOW_COLORS[us["moscow"]])
        mc.font = Font(bold=True, color="FFFFFF")

        # MVP coloration
        mvp_cell = ws.cell(row=r, column=9)
        if us["mvp"] == "Oui":
            mvp_cell.fill = PatternFill("solid", fgColor="C6EFCE")
            mvp_cell.font = Font(bold=True, color="006100")

        ws.row_dimensions[r].height = 90

    # Total ligne
    total_row = len(US) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    col_charge = 11
    ws.cell(row=total_row, column=col_charge,
            value=f"=SUM({get_column_letter(col_charge)}2:{get_column_letter(col_charge)}{total_row-1})").font = Font(bold=True)
    for i, _ in enumerate(PROFILS):
        col = col_charge + 1 + i
        ws.cell(row=total_row, column=col,
                value=f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{total_row-1})").font = Font(bold=True)
    cost_col = col_charge + 1 + len(PROFILS)
    ws.cell(row=total_row, column=cost_col,
            value=f"=SUM({get_column_letter(cost_col)}2:{get_column_letter(cost_col)}{total_row-1})").font = Font(bold=True)
    for c in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=c).fill = PatternFill("solid", fgColor="DDEBF7")

    # largeur colonnes
    widths = {1: 8, 2: 38, 3: 22, 4: 38, 5: 32, 6: 8, 7: 10, 8: 8, 9: 6, 10: 32, 11: 10}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    for i in range(len(PROFILS)):
        ws.column_dimensions[get_column_letter(12 + i)].width = 14
    ws.column_dimensions[get_column_letter(12 + len(PROFILS))].width = 14
    ws.freeze_panes = "B2"

    # ---------- Onglet Synthèse coûts ----------
    if "Synthèse coûts" in wb.sheetnames:
        del wb["Synthèse coûts"]
    s = wb.create_sheet("Synthèse coûts")
    s.append(["Profil", "TJM (€)", "Total jours", "Coût (€)"])
    style_header_row(s, 1, 4)
    for prof, tjm in PROFILS:
        total_j = sum(round(us["charge"] * us["rep"].get(prof, 0) / 100, 2) for us in US)
        s.append([prof, tjm, round(total_j, 2), round(total_j * tjm, 0)])
    last = s.max_row
    s.append(["TOTAL", "", f"=SUM(C2:C{last})", f"=SUM(D2:D{last})"])
    for c in range(1, 5):
        s.cell(row=s.max_row, column=c).font = Font(bold=True)
        s.cell(row=s.max_row, column=c).fill = PatternFill("solid", fgColor="DDEBF7")
    s.column_dimensions["A"].width = 40
    s.column_dimensions["B"].width = 12
    s.column_dimensions["C"].width = 14
    s.column_dimensions["D"].width = 16

    # ---------- Onglet Planning sprints ----------
    if "Planning sprints" in wb.sheetnames:
        del wb["Planning sprints"]
    p = wb.create_sheet("Planning sprints")
    p.append(["Sprint", "Thème", "User stories", "Charge totale (j)", "Livrable / Objectif"])
    style_header_row(p, 1, 5)
    themes = {
        1: ("Sprint 1 — Socle utilisateur & ingestion", "Compte utilisateur opérationnel + pipeline d'ingestion des photos"),
        2: ("Sprint 2 — Gestion garde-robe & 1er modèle IA", "Garde-robe gérable + détection vêtements (PoC vision)"),
        3: ("Sprint 3 — Moteur de recommandation & tunnel d'achat (MVP complet)", "MVP des deux moteurs IA (garde-robe et préférences/tendances) + panier/commande/paiement : la boucle recommandation → achat est mesurable dès ce sprint"),
        4: ("Sprint 4 — RGPD & feedback", "Conformité RGPD validée par le DPO + boucle de feedback"),
        5: ("Sprint 5 — Personnalisation & purge", "Personnalisation visuelle + purge automatique"),
        6: ("Sprint 6 — Sources externes & polissage", "Intégration sources externes + recettes finales"),
    }
    for sp_num in sorted(themes.keys()):
        us_in_sprint = [u for u in US if u["sprint"] == sp_num]
        charge = sum(u["charge"] for u in us_in_sprint)
        ids = ", ".join(u["id"] for u in us_in_sprint)
        p.append([f"Sprint {sp_num}", themes[sp_num][0], ids, charge, themes[sp_num][1]])
    p.column_dimensions["A"].width = 10
    p.column_dimensions["B"].width = 42
    p.column_dimensions["C"].width = 28
    p.column_dimensions["D"].width = 16
    p.column_dimensions["E"].width = 50
    for row in p.iter_rows(min_row=2, values_only=False):
        for cell in row:
            cell.alignment = LEFT_WRAP
            cell.border = BORDER

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Saved {OUT}")


if __name__ == "__main__":
    build()
