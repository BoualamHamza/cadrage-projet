"""Génère Chiffrage_Couts_Rentabilite.xlsx.

Onglets :
- Hypothèses : TJM, coûts Azure, gains marketing, taux maintenance
- Coût développement initial : agrégation depuis le backlog
- Coût Azure (initial + production) : estimation
- Rentabilité : tableau coûts/gains cumulés année par année + graphique
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "livrables" / "Chiffrage_Couts_Rentabilite.xlsx"
BACKLOG = ROOT / "livrables" / "Backlog_Fashion_Insta.xlsx"

HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
SECT_FILL = PatternFill("solid", fgColor="2E75B6")
SECT_FONT = Font(bold=True, color="FFFFFF", size=12)
TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def hdr(ws, row, values):
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def section(ws, row, label, ncols=4):
    ws.cell(row=row, column=1, value=label)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = SECT_FILL
        cell.font = SECT_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def build():
    wb = openpyxl.Workbook()

    # ====== Onglet 1 : Hypothèses ======
    h = wb.active
    h.title = "Hypothèses"
    h.column_dimensions["A"].width = 50
    h.column_dimensions["B"].width = 18
    h.column_dimensions["C"].width = 40

    section(h, 1, "Hypothèses générales", ncols=3)
    h.append(["Durée d'un sprint (semaines)", "2 (Sprint 3 : 3)", "Sprint 3 étendu à 3 semaines : 60 j de charge ne tiennent pas sur 10 j ouvrés (2 sem.) sans mobiliser 6 personnes à temps plein en simultané"])
    h.append(["Nombre de sprints", 6, "Couvre l'ensemble du backlog priorisé"])
    h.append(["Durée totale du projet (semaines)", 13, "5 sprints × 2 semaines + Sprint 3 × 3 semaines"])
    h.append(["Taux de maintenance annuel", 0.15, "15% du coût de développement initial (règle PO)"])
    h.append(["Horizon d'analyse rentabilité (années)", 5, "Standard projet IA"])

    h.append([])
    section(h, h.max_row + 1, "TJM par profil (€/jour)", ncols=3)
    tjms = [
        ("PO / Product Manager", 600),
        ("Tech Lead", 800),
        ("Data Scientist junior", 500),
        ("Data Scientist senior (sous-traitant)", 1000),
        ("ML / Computer Vision Engineer", 700),
        ("Développeur mobile", 600),
        ("DevOps / Cloud Azure", 700),
        ("UX / UI Designer", 550),
        ("QA / Testeur", 500),
    ]
    for label, val in tjms:
        h.append([label, val, "TJM marché France 2026"])

    h.append([])
    section(h, h.max_row + 1, "Coûts Azure", ncols=3)
    azure_init = [
        ("Stockage Blob (jeux d'entraînement, images)", 1200, "Phase entraînement, ~3 mois"),
        ("Azure Machine Learning (compute GPU entraînement, NC6s v3)", 8000, "≈ 2000h GPU cumulées sur 3 mois (usage quasi continu : itérations + hyperparameter tuning). NC6s v3 ≈ 3,8 €/h — le SKU 'NC6' d'origine est retiré du catalogue Azure"),
        ("Azure Custom Vision / Cognitive Services (annotation)", 1500, "Pour les itérations rapides"),
        ("Environnements de dev / pré-prod", 2000, "App Service + DB managée"),
        ("Sécurité & monitoring (Microsoft Defender for Cloud)", 1300, "Sur la phase projet, sans SIEM Sentinel (coût prohibitif pour ce périmètre, cf. onglet prod)"),
    ]
    for label, val, desc in azure_init:
        h.append([label, val, desc])
    h.cell(row=h.max_row + 1, column=1, value="Total Azure initial (€)").font = Font(bold=True)
    h.cell(row=h.max_row, column=2, value=sum(a[1] for a in azure_init)).font = Font(bold=True)

    h.append([])
    section(h, h.max_row + 1, "Coûts Azure annuels (production)", ncols=3)
    # Postes "variables" : évoluent avec le nombre d'utilisateurs actifs (hébergement, inférence, stockage)
    # Postes "fixes" : n'évoluent pas avec l'usage (base de données, sécurité) — approximation conservatrice
    azure_prod = [
        ("Hébergement API recommandation (Azure App Service, auto-scaling)", 12000, "Variable — charge moyenne à 50k utilisateurs actifs (An 1)"),
        ("Inférence ML (Azure ML Endpoint, CPU + cache embeddings, scale-to-zero)", 14000, "Variable — ≈ 1M requêtes/mois à 50k utilisateurs actifs (An 1). Suppose une inférence CPU (embeddings pré-calculés) ; si un modèle GPU temps réel s'avère nécessaire, ce poste serait à revoir nettement à la hausse (un seul GPU NC6s v3 24/7 coûte à lui seul ≈ 31 000€/an)"),
        ("Stockage Blob production (photos utilisateurs)", 6000, "Variable — chiffré, redondance ZRS, à 50k utilisateurs actifs (An 1)"),
        ("Base de données managée (Azure SQL / Cosmos DB)", 5000, "Fixe — compte utilisateur + préférences"),
        ("Sécurité, sauvegarde, monitoring (Defender + Azure Monitor)", 2000, "Fixe — sans SIEM Sentinel : à 5,59 $/Go analysé, Sentinel dépasserait ce budget avec quelques dizaines de Go/mois ingérés"),
    ]
    N_AZURE_VARIABLE = 3  # les 3 premiers postes ci-dessus sont variables, les 2 derniers sont fixes
    for label, val, desc in azure_prod:
        h.append([label, val, desc])
    h.cell(row=h.max_row + 1, column=1, value="Total Azure annuel (€, base An 1 à 50k utilisateurs)").font = Font(bold=True)
    h.cell(row=h.max_row, column=2, value=sum(a[1] for a in azure_prod)).font = Font(bold=True)
    h.append(["Note : les postes variables sont indexés sur √(utilisateurs actifs / 50 000) d'une année sur l'autre", "", "Hypothèse d'économies d'échelle — évite une croissance linéaire irréaliste des coûts cloud ; détail par année dans l'onglet Rentabilité"])

    h.append([])
    section(h, h.max_row + 1, "Hypothèses gains marketing", ncols=3)
    gains = [
        ("Nb utilisateurs actifs an 1", 50000, "Lancement progressif"),
        ("Nb utilisateurs actifs an 2", 150000, "Marketing & bouche à oreille"),
        ("Nb utilisateurs actifs an 3", 280000, "Croissance maintenue"),
        ("Nb utilisateurs actifs an 4", 380000, "Plafonnement"),
        ("Nb utilisateurs actifs an 5", 450000, "Maturité"),
        ("Taux de conversion recommandation → achat", 0.08, "Part des utilisateurs actifs générant réellement un panier additionnel via la reco (aligné avec le KPI cible ≥5%, cf. présentation COMEX) — un utilisateur actif n'est pas un acheteur"),
        ("Panier additionnel moyen / utilisateur convertissant / an (€)", 25, "Upsell estimé par le marketing, appliqué uniquement aux utilisateurs qui convertissent"),
        ("Marge brute retail moyenne", 0.45, "Pour convertir le CA additionnel en gain net comparable aux coûts (le panier ci-dessus est du CA, pas du profit)"),
    ]
    for label, val, desc in gains:
        h.append([label, val, desc])

    azure_initial_total = sum(a[1] for a in azure_init)
    azure_annuel_total = sum(a[1] for a in azure_prod)

    # ====== Onglet 2 : Coût développement initial ======
    c = wb.create_sheet("Coût développement")
    c.column_dimensions["A"].width = 45
    c.column_dimensions["B"].width = 14
    c.column_dimensions["C"].width = 14
    c.column_dimensions["D"].width = 16

    hdr(c, 1, ["Profil", "TJM (€)", "Jours total", "Coût (€)"])
    # Récup des totaux par profil depuis le backlog
    wb_b = openpyxl.load_workbook(BACKLOG, data_only=False)
    s_b = wb_b["Synthèse coûts"]
    rows = []
    for row in s_b.iter_rows(min_row=2, values_only=True):
        if row[0] and row[0] != "TOTAL":
            rows.append((row[0], row[1], row[2], row[3]))

    total_jours = 0
    total_cost = 0
    for r, (prof, tjm, jrs, cost) in enumerate(rows, start=2):
        c.cell(row=r, column=1, value=prof).alignment = LEFT_WRAP
        c.cell(row=r, column=2, value=tjm).alignment = CENTER
        c.cell(row=r, column=3, value=jrs).alignment = CENTER
        c.cell(row=r, column=4, value=cost).alignment = CENTER
        total_jours += jrs if isinstance(jrs, (int, float)) else 0
        total_cost += cost if isinstance(cost, (int, float)) else 0
        for col in range(1, 5):
            c.cell(row=r, column=col).border = BORDER
    tr = c.max_row + 1
    c.cell(row=tr, column=1, value="TOTAL équipe").font = Font(bold=True)
    c.cell(row=tr, column=3, value=total_jours).font = Font(bold=True)
    c.cell(row=tr, column=4, value=total_cost).font = Font(bold=True)
    for col in range(1, 5):
        c.cell(row=tr, column=col).fill = TOTAL_FILL
        c.cell(row=tr, column=col).border = BORDER

    c.append([])
    c.append(["Coût RH développement (€)", "", "", total_cost])
    c.append(["+ Coût Azure initial (€)", "", "", azure_initial_total])
    grand_total = total_cost + azure_initial_total
    c.append(["= COÛT INITIAL PROJET (€)", "", "", grand_total])
    for r in range(c.max_row - 2, c.max_row + 1):
        for col in range(1, 5):
            c.cell(row=r, column=col).font = Font(bold=True)
            c.cell(row=r, column=col).fill = TOTAL_FILL
            c.cell(row=r, column=col).border = BORDER

    # ====== Onglet 3 : Coûts récurrents ======
    rec = wb.create_sheet("Coûts récurrents")
    rec.column_dimensions["A"].width = 55
    rec.column_dimensions["B"].width = 18

    hdr(rec, 1, ["Poste", "Montant annuel (€, base An 1)", "Type"])
    rec.column_dimensions["C"].width = 22
    maintenance = round(total_cost * 0.15, 0)
    azure_variable_base = sum(a[1] for a in azure_prod[:N_AZURE_VARIABLE])
    azure_fixed = sum(a[1] for a in azure_prod[N_AZURE_VARIABLE:])
    items = [
        ("Maintenance applicative (15% du coût RH initial)", maintenance, "Fixe"),
        ("Infrastructure Azure — hébergement/inférence/stockage", azure_variable_base, "Variable (usagers)"),
        ("Infrastructure Azure — BDD, sécurité & monitoring", azure_fixed, "Fixe"),
        ("Ré-entraînement modèles (compute + DS junior 20j)", 18000, "Fixe"),
        ("DPO / RGPD - veille et audits", 6000, "Fixe"),
    ]
    for r, (label, val, typ) in enumerate(items, start=2):
        rec.cell(row=r, column=1, value=label).alignment = LEFT_WRAP
        rec.cell(row=r, column=2, value=val).alignment = CENTER
        rec.cell(row=r, column=3, value=typ).alignment = CENTER
        for col in range(1, 4):
            rec.cell(row=r, column=col).border = BORDER
    tr = rec.max_row + 1
    rec.cell(row=tr, column=1, value="TOTAL annuel (€, base An 1 à 50k utilisateurs)").font = Font(bold=True)
    annuel_total = sum(i[1] for i in items)
    fixed_recurring_total = maintenance + azure_fixed + 18000 + 6000
    rec.cell(row=tr, column=2, value=annuel_total).font = Font(bold=True)
    for col in range(1, 4):
        rec.cell(row=tr, column=col).fill = TOTAL_FILL
        rec.cell(row=tr, column=col).border = BORDER
    rec.cell(row=tr + 2, column=1, value=(
        "Le poste 'Variable (usagers)' croît avec le nombre d'utilisateurs actifs (facteur "
        "√(utilisateurs / 50 000)) — voir la progression année par année dans l'onglet Rentabilité."
    )).alignment = LEFT_WRAP
    rec.cell(row=tr + 2, column=1).font = Font(italic=True)
    rec.merge_cells(start_row=tr + 2, start_column=1, end_row=tr + 2, end_column=3)
    rec.row_dimensions[tr + 2].height = 30

    # ====== Onglet 4 : Rentabilité ======
    r_ws = wb.create_sheet("Rentabilité")
    r_ws.column_dimensions["A"].width = 36
    for col in "BCDEFG":
        r_ws.column_dimensions[col].width = 14

    hdr(r_ws, 1, ["Poste (€)", "Année 0", "Année 1", "Année 2", "Année 3", "Année 4", "Année 5"])

    users = [50000, 150000, 280000, 380000, 450000]
    taux_conversion = 0.08
    panier_an = 25
    marge = 0.45
    # Gain net = utilisateurs actifs x taux de conversion x panier additionnel x marge brute
    # (et non "utilisateurs actifs x panier", qui supposerait que 100% des actifs achètent)
    gains_annuels = [round(u * taux_conversion * panier_an * marge, 0) for u in users]

    # Coûts : le socle fixe reste constant, la part variable (hébergement/inférence/stockage)
    # est indexée sur √(utilisateurs / 50 000) pour refléter la croissance de l'usage sans
    # supposer une croissance linéaire irréaliste des coûts cloud (économies d'échelle).
    users_ref = users[0]
    couts_par_annee = [grand_total] + [
        round(fixed_recurring_total + azure_variable_base * (u / users_ref) ** 0.5, 0)
        for u in users
    ]
    r_ws.append(["Coûts (initial / annuels)"] + couts_par_annee)
    r_ws.append(["Coûts cumulés"] + [None] * 6)
    r_ws.append(["Gains annuels"] + [0] + gains_annuels)
    r_ws.append(["Gains cumulés"] + [None] * 6)
    r_ws.append(["Solde cumulé (gains - coûts)"] + [None] * 6)

    # remplir avec formules
    # rangée 2 = Coûts annuels, 3 = Coûts cumulés, 4 = Gains annuels, 5 = Gains cumulés, 6 = Solde
    for col in range(2, 8):  # B..G
        L = get_column_letter(col)
        if col == 2:
            r_ws[f"{L}3"] = f"={L}2"
            r_ws[f"{L}5"] = f"={L}4"
        else:
            prev = get_column_letter(col - 1)
            r_ws[f"{L}3"] = f"={prev}3+{L}2"
            r_ws[f"{L}5"] = f"={prev}5+{L}4"
        r_ws[f"{L}6"] = f"={L}5-{L}3"

    # formatting
    for r in range(2, 7):
        for col in range(1, 8):
            cell = r_ws.cell(row=r, column=col)
            cell.border = BORDER
            if col == 1:
                cell.font = Font(bold=True)
            else:
                cell.number_format = "#,##0 €"
        r_ws.cell(row=r, column=1).alignment = LEFT_WRAP

    # surligner la ligne solde cumulé
    for col in range(1, 8):
        r_ws.cell(row=6, column=col).fill = TOTAL_FILL
        r_ws.cell(row=6, column=col).font = Font(bold=True)

    # Graphique
    chart = LineChart()
    chart.title = "Rentabilité — coûts vs gains cumulés"
    chart.y_axis.title = "€"
    chart.x_axis.title = "Année"
    chart.height = 10
    chart.width = 22
    data = Reference(r_ws, min_col=1, min_row=3, max_row=6, max_col=8)
    chart.add_data(data, titles_from_data=True, from_rows=True)
    cats = Reference(r_ws, min_col=2, min_row=1, max_col=8, max_row=1)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList(showVal=False)
    r_ws.add_chart(chart, "A10")

    # Commentaire de lecture
    r_ws["A28"] = "Lecture : le projet devient rentable l'année où le 'Solde cumulé' devient positif."
    r_ws["A28"].font = Font(italic=True)
    r_ws["A29"] = (
        f"Gains = utilisateurs actifs × taux de conversion ({taux_conversion:.0%}) × panier additionnel "
        f"({panier_an}€) × marge brute ({marge:.0%}) — pas 100% des utilisateurs actifs, et en gain net, pas en CA brut."
    )
    r_ws["A29"].font = Font(italic=True)
    r_ws["A30"] = "Coûts annuels = socle fixe + coûts Azure variables indexés sur √(utilisateurs actifs / 50 000)."
    r_ws["A30"].font = Font(italic=True)

    # Sauvegarde
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Saved {OUT}")
    print(f"Coût initial: {grand_total} €, annuel An 1: {annuel_total} €")
    print(f"Coûts par année (0-5): {couts_par_annee}")
    print(f"Gains an1-5: {gains_annuels}")


if __name__ == "__main__":
    build()
