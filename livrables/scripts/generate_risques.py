"""Génère Analyse_des_risques.xlsx à partir du template + checklist spectre 7D.

7D = Données / Délais / Dépendances / Dimensionnement (ressources) /
     Décisions (gouvernance) / Déontologie (éthique/RGPD) / Déploiement.
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

ROOT = Path(__file__).resolve().parents[2]
TEMPLATE = ROOT / "pieces_jointes" / "Modèle_analyse_des_risques.xlsx"
OUT = ROOT / "livrables" / "Analyse_des_risques.xlsx"

HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Lignes (axe 7D, facteur, risque, conséquence, conséquences détaillées,
#         impact, proba, actions prévention, actions correction)
RISQUES = [
    ("Données",
     "Les images des utilisateurs sont des données personnelles sensibles",
     "Fuite ou accès non autorisé aux photos utilisateurs",
     "Atteinte à la vie privée, plainte CNIL, perte de confiance",
     "Coût : amende RGPD jusqu'à 4% CA. Qualité : perte de confiance. Satisfaction : très forte chute.",
     3, 2,
     "Chiffrement AES-256, accès restreint par rôle, MFA, audit annuel, AIPD avant prod, pen-test régulier",
     "Activation du plan de réponse incident, notification CNIL sous 72h, communication aux utilisateurs, purge des données compromises"),

    ("Données",
     "Datasets publics d'entraînement peu diversifiés (DeepFashion etc.)",
     "Modèle biaisé envers certaines morphologies / origines / genres",
     "Recommandations peu pertinentes pour certains profils, atteinte à l'image, risque discriminatoire",
     "Qualité : KPI dégradés sur certains segments. Satisfaction : abandon utilisateur. Risque réputationnel.",
     3, 3,
     "Audit du dataset, métriques d'équité par sous-groupe, enrichissement avec données diversifiées, revue éthique du modèle avant release",
     "Bridage du modèle sur les segments à risque, ré-entraînement, communication transparente"),

    ("Délais",
     "Concurrent prépare une application similaire — pression time-to-market",
     "Glissement du planning et lancement après le concurrent",
     "Perte d'avantage concurrentiel, baisse des gains attendus",
     "Coût : ROI repoussé. Délai : perte de parts de marché.",
     2, 3,
     "Priorisation MoSCoW stricte, MVP centré sur l'IA core, planning en sprints courts (2 semaines), buffer de 20% sur les sprints critiques",
     "Découpage du périmètre, sortie d'un MVP réduit avant montée en charge"),

    ("Dépendances",
     "Développeurs mobiles en parallèle sur une autre application urgente",
     "Sous-disponibilité de l'équipe mobile, retards sur les US client",
     "Retard global, frustration équipe, perte de qualité",
     "Délai : décalage de 2-3 sprints. Qualité : code rushé.",
     2, 3,
     "Engagement formel sur capacité de l'équipe mobile, plan de charge partagé, escalade hebdo au comité de pilotage",
     "Recours à un développeur mobile freelance, replanification des US non bloquantes"),

    ("Dimensionnement",
     "Data Scientist junior en interne, expertise CV limitée",
     "Modèle IA en deçà des attentes en termes de performance",
     "Recommandations peu pertinentes, désengagement utilisateurs",
     "Qualité : modèle médiocre. Satisfaction : faible adoption.",
     3, 2,
     "Sous-traitance avec DS senior cadrée par contrat, pair-coding hebdo, peer-review systématique, transfert de compétence planifié",
     "Renforcement du sous-traitant, prolongation de la phase d'entraînement, simplification du modèle"),

    ("Décisions",
     "Forte pression de la direction pour aller vite, gouvernance projet à roder",
     "Décisions prises sans validation DPO ou design éthique",
     "Non-conformité RGPD, dette éthique",
     "Coût : refonte forcée. Qualité : confiance dégradée.",
     2, 2,
     "Comité de pilotage bi-mensuel avec DPO + VP Product, definition of done incluant validation DPO pour les US IA",
     "Pause de la US concernée, revue éthique extraordinaire, re-cadrage"),

    ("Déontologie",
     "Recommandations IA basées sur photo : sujet sensible (image de soi, biais)",
     "Polémique éthique ou médiatique (presse, réseaux sociaux)",
     "Risque réputationnel, abandon utilisateurs, audit imposé",
     "Satisfaction : forte baisse. Coût : audit / refonte. Qualité : perte de crédibilité.",
     3, 2,
     "Charte éthique IA publiée, transparence sur l'usage IA, opt-in clair, KPI d'équité publiés, comité éthique interne",
     "Communication de crise pré-rédigée, retrait temporaire de la fonctionnalité, audit indépendant"),

    ("Déontologie",
     "Photos = données sensibles — obligation de minimisation",
     "Conservation des photos au-delà de la durée légitime",
     "Non-conformité RGPD, sanction CNIL",
     "Coût : amende. Qualité : audit imposé.",
     2, 2,
     "Purge automatique (US15), durée de conservation 12 mois max, monitoring du job de purge",
     "Purge manuelle d'urgence, notification CNIL si besoin"),

    ("Déploiement",
     "Coûts Azure d'inférence sous-estimés à grande échelle",
     "Dépassement budgétaire en production",
     "Rentabilité repoussée, arbitrage difficile",
     "Coût : surcoût annuel pouvant doubler le budget Azure.",
     2, 2,
     "Tests de charge représentatifs, benchmark Azure Calculator, alerting budgétaire, optimisation modèle (quantization, distillation)",
     "Optimisation du modèle (quantization, cache), passage à des SKUs moins coûteux, négociation Microsoft"),

    ("Déploiement",
     "Lancement mobile sur deux stores (iOS + Android) avec validations externes",
     "Refus / délai de validation par Apple ou Google",
     "Lancement repoussé, perte d'avantage concurrentiel",
     "Délai : décalage de plusieurs semaines.",
     2, 2,
     "Soumission anticipée, conformité aux guidelines, version web PWA en fallback",
     "Itération sur les remarques des stores, communication contournée vers le canal web"),

    ("Données",
     "Volumétrie d'images en croissance rapide",
     "Saturation stockage / coûts inattendus",
     "Coûts Azure qui dérivent, performance dégradée",
     "Coût : dérive du budget Azure. Qualité : latence accrue.",
     2, 2,
     "Compression intelligente, archivage tiered (hot/cool/cold), monitoring volumétrie",
     "Rétention raccourcie sur ancien historique, plan de purge ciblé"),

    ("Dimensionnement",
     "Sous-traitant DS senior critique pour le modèle",
     "Indisponibilité du sous-traitant en cours de projet",
     "Blocage de l'avancement IA, perte de connaissance",
     "Délai : décalage. Qualité : régression modèle.",
     3, 1,
     "Contrat avec engagement de continuité, documentation des modèles versionnée, pair-coding systématique DS junior",
     "Activation d'un sous-traitant de back-up, prolongation de phase, mode dégradé sur le moteur"),
]


def build():
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb["Risques"]
    ws.delete_rows(1, ws.max_row)

    headers = [
        "Axe 7D",
        "Facteur de risque (étant donné que...)",
        "Risque — événement redouté (si...)",
        "Conséquence",
        "Conséquences détaillées (coût/délai/qualité/satisfaction) (alors...)",
        "Impact (0-3)",
        "Probabilité (0-3)",
        "Criticité",
        "Actions de prévention",
        "Actions de correction",
    ]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[1].height = 45

    for r, (axe, fact, risque, cons, cons_detail, impact, proba, prev, corr) in enumerate(RISQUES, start=2):
        ws.cell(row=r, column=1, value=axe)
        ws.cell(row=r, column=2, value=fact)
        ws.cell(row=r, column=3, value=risque)
        ws.cell(row=r, column=4, value=cons)
        ws.cell(row=r, column=5, value=cons_detail)
        ws.cell(row=r, column=6, value=impact)
        ws.cell(row=r, column=7, value=proba)
        ws.cell(row=r, column=8, value=impact * proba)
        ws.cell(row=r, column=9, value=prev)
        ws.cell(row=r, column=10, value=corr)
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = LEFT_WRAP if c in (2, 3, 4, 5, 9, 10) else CENTER
        ws.row_dimensions[r].height = 110

    # Mise en couleur de la criticité
    last_row = ws.max_row
    rule = ColorScaleRule(
        start_type="num", start_value=0, start_color="63BE7B",
        mid_type="num", mid_value=4, mid_color="FFEB84",
        end_type="num", end_value=9, end_color="F8696B",
    )
    ws.conditional_formatting.add(f"H2:H{last_row}", rule)

    widths = {1: 16, 2: 36, 3: 34, 4: 32, 5: 38, 6: 8, 7: 9, 8: 9, 9: 38, 10: 36}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "B2"

    # ====== Onglet 'Spectre 7D' synthèse ======
    if "Spectre 7D" in wb.sheetnames:
        del wb["Spectre 7D"]
    s = wb.create_sheet("Spectre 7D")
    s.column_dimensions["A"].width = 24
    s.column_dimensions["B"].width = 70
    s.column_dimensions["C"].width = 12

    s.append(["Axe", "Définition / périmètre", "Nb risques"])
    for c in range(1, 4):
        cell = s.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER

    axes_def = [
        ("Données", "Qualité, accès, sécurité, sensibilité, biais des datasets"),
        ("Délais", "Pression temporelle, fenêtre concurrentielle, planning"),
        ("Dépendances", "Adhérences à d'autres équipes, fournisseurs, sous-traitants"),
        ("Dimensionnement", "Compétences et disponibilité de l'équipe, profils critiques"),
        ("Décisions", "Gouvernance, escalade, arbitrage produit / sécurité / éthique"),
        ("Déontologie", "Éthique IA, RGPD, biais, transparence"),
        ("Déploiement", "Mise en production, infra, distribution mobile, montée en charge"),
    ]
    for axe, defi in axes_def:
        count = sum(1 for r in RISQUES if r[0] == axe)
        s.append([axe, defi, count])
    for r in range(2, s.max_row + 1):
        for c in range(1, 4):
            cell = s.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = LEFT_WRAP if c == 2 else CENTER

    # ====== Onglet 'Top risques' ======
    if "Top risques" in wb.sheetnames:
        del wb["Top risques"]
    t = wb.create_sheet("Top risques")
    t.append(["Rang", "Axe", "Risque", "Criticité", "Actions clés"])
    for c in range(1, 6):
        cell = t.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
    sorted_r = sorted(RISQUES, key=lambda x: x[5] * x[6], reverse=True)
    for rank, (axe, fact, risque, cons, cd, imp, prob, prev, corr) in enumerate(sorted_r[:5], start=1):
        t.append([rank, axe, risque, imp * prob, prev.split(",")[0]])
    for r in range(2, t.max_row + 1):
        for c in range(1, 6):
            cell = t.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = LEFT_WRAP if c in (3, 5) else CENTER
    t.column_dimensions["A"].width = 6
    t.column_dimensions["B"].width = 16
    t.column_dimensions["C"].width = 60
    t.column_dimensions["D"].width = 10
    t.column_dimensions["E"].width = 60

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Saved {OUT}")
    print(f"Nb risques : {len(RISQUES)}")


if __name__ == "__main__":
    build()
